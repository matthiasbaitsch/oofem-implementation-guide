import openpyxl

from structure import Structure

def read_excel(file: str) -> Structure:
    workbook = openpyxl.load_workbook(file, data_only=True)
    structure = Structure()

    # Nodes
    nodes_sheet = workbook.worksheets[0]

    for row in nodes_sheet.iter_rows(min_row=2, values_only=True):
        node = structure.add_node(row[0], row[1], row[2])
        node.constraint.fixed[0] = bool(row[3])
        node.constraint.fixed[1] = bool(row[4])
        node.constraint.fixed[2] = bool(row[5])
        node.force.components[0] = row[6] or 0.0
        node.force.components[1] = row[7] or 0.0
        node.force.components[2] = row[8] or 0.0

    # Elements
    elements_sheet = workbook.worksheets[1]
    for row in elements_sheet.iter_rows(min_row=2, values_only=True):
        structure.add_element(row[0], row[1], int(row[2]), int(row[3]))

    return structure
